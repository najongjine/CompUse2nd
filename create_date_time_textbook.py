from copy import copy
from datetime import datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
TEXTBOOK_DIR = next(ROOT.rglob("2_논리함수.xlsx")).parent
OUTPUT = TEXTBOOK_DIR / "2_날짜와 시간 함수.xlsx"

NAVY = "17365D"
BLUE = "DDEBF7"
BLUE_DARK = "5B9BD5"
GREEN = "E2F0D9"
GREEN_DARK = "70AD47"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
PURPLE = "E4DFEC"
GRAY = "F2F2F2"
WHITE = "FFFFFF"
RED = "F4CCCC"
BLACK = "000000"
THIN_GRAY = Side(style="thin", color="B7B7B7")
MEDIUM_NAVY = Side(style="medium", color=NAVY)


def fill(color):
    return PatternFill("solid", fgColor=color)


def setup_sheet(ws, widths=None, freeze="A4"):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    widths = widths or [14, 18, 16, 16, 18, 18, 18, 18, 18, 18, 18, 18]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.outlinePr.summaryBelow = True


def merge_write(ws, cell_range, value, color=None, font=None, align=None, border=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if color:
        cell.fill = fill(color)
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        for row in ws[cell_range]:
            for item in row:
                item.border = border
    return cell


def title(ws, heading, subtitle, end_col=10):
    merge_write(
        ws,
        f"A1:{get_column_letter(end_col)}1",
        heading,
        NAVY,
        Font(name="맑은 고딕", size=16, bold=True, color=WHITE),
        Alignment(vertical="center"),
    )
    ws.row_dimensions[1].height = 28
    merge_write(
        ws,
        f"A2:{get_column_letter(end_col)}2",
        subtitle,
        BLUE,
        Font(name="맑은 고딕", size=10, color=NAVY),
        Alignment(wrap_text=True, vertical="center"),
    )
    ws.row_dimensions[2].height = 28


def section(ws, row, text, end_col=10, color=NAVY):
    merge_write(
        ws,
        f"A{row}:{get_column_letter(end_col)}{row}",
        text,
        color,
        Font(name="맑은 고딕", size=11, bold=True, color=WHITE),
        Alignment(vertical="center"),
    )
    ws.row_dimensions[row].height = 22


def table(ws, start_row, headers, rows, formats=None, fills=None):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = fill(NAVY)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=MEDIUM_NAVY, bottom=MEDIUM_NAVY, left=THIN_GRAY, right=THIN_GRAY)
    for r_offset, values in enumerate(rows, 1):
        for col, value in enumerate(values, 1):
            cell = ws.cell(start_row + r_offset, col, value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
            if fills and col in fills:
                cell.fill = fill(fills[col])
            elif r_offset % 2 == 0:
                cell.fill = fill("F8FBFF")
            if formats and col in formats:
                cell.number_format = formats[col]
    return start_row + len(rows)


def note(ws, cell_range, text, color=YELLOW):
    merge_write(
        ws,
        cell_range,
        text,
        color,
        Font(name="맑은 고딕", size=10, bold=True, color=NAVY),
        Alignment(wrap_text=True, vertical="center"),
        Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY),
    )


def style_all(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and not cell.font.name:
                cell.font = Font(name="맑은 고딕", size=10)


def make_cover(wb):
    ws = wb.active
    ws.title = "표지_4시간계획"
    setup_sheet(ws, [16, 22, 46, 22, 22, 22], None)
    merge_write(
        ws,
        "A1:F1",
        "컴활 2급 실기 - 날짜와 시간 함수 4시간 학습지",
        NAVY,
        Font(name="맑은 고딕", size=17, bold=True, color=WHITE),
        Alignment(vertical="center"),
    )
    ws.row_dimensions[1].height = 30
    merge_write(
        ws,
        "A2:F2",
        "컴맹 기준: 달력으로 이해 → 함수 조각 보기 → 따라 쓰기 → 빈칸 연습 → 시험형 문제 → 오답 교정",
        BLUE,
        Font(name="맑은 고딕", size=10, color=NAVY),
        Alignment(wrap_text=True, vertical="center"),
    )
    ws.row_dimensions[2].height = 28
    note(
        ws,
        "A4:F5",
        "이 교재의 목표\n날짜 함수 이름을 외우는 것이 아닙니다. 문제 문장에서 '꺼낼 것', '만들 것', '간격을 잴 것'을 찾아 알맞은 함수를 고르는 연습을 합니다.",
    )
    table(
        ws,
        7,
        ["시간", "시트", "배우는 것", "학생 활동", "시험 연결", "완료"],
        [
            ["0:00-0:20", "0_날짜는숫자", "엑셀이 날짜와 시간을 저장하는 방식", "달력 타임라인 보기", "날짜 계산 실수 예방", "□"],
            ["0:20-0:40", "0_기간계산공식", "근속일수, 근무시간, 경과일 비교", "단위에 따라 공식 고르기", "기간 계산", "□"],
            ["0:40-1:15", "1교시_오늘과추출", "TODAY, NOW, YEAR, MONTH, DAY", "생년월일/입사일 분해", "나이, 근속연수", "□"],
            ["1:15-1:55", "2교시_만들기와시간", "DATE, TIME, HOUR, MINUTE, SECOND", "조각을 합쳐 날짜/시간 만들기", "만기일, 근무시간", "□"],
            ["1:55-2:35", "3교시_요일과기간", "WEEKDAY, 날짜 빼기", "요일표와 기간 계산", "요일, 경과일", "□"],
            ["2:35-3:15", "4교시_납기와월말", "WORKDAY, EDATE, EOMONTH", "납기일/월말 계산", "근무일, 만기일", "□"],
            ["3:15-4:00", "실전연습", "컴활 2급 실기형 문제", "초록칸 직접 입력", "혼합 출제 대응", "□"],
        ],
        fills={6: GREEN},
    )
    note(ws, "A16:F17", "색깔 약속: 파란칸 = 설명용 완성 예제 / 초록칸 = 직접 수식 입력 / 노란칸 = 시험에서 잡아야 할 단서 / 보라칸 = 정답 확인")
    ws.row_dimensions[16].height = 34


def make_serial_sheet(wb):
    ws = wb.create_sheet("0_날짜는숫자")
    setup_sheet(ws, [18, 18, 18, 18, 18, 18, 18, 18, 18, 18])
    title(ws, "0. 날짜는 숫자다", "날짜 계산이 되는 이유부터 눈으로 봅니다. 이 원리를 알면 함수가 덜 무섭습니다.")
    section(ws, 4, "그림 1. 엑셀 달력은 숫자 자처럼 이어져 있다")
    dates = [datetime(2026, 5, 30), datetime(2026, 5, 31), datetime(2026, 6, 1), datetime(2026, 6, 2), datetime(2026, 6, 3)]
    serials = [d.toordinal() - datetime(1899, 12, 30).toordinal() for d in dates]
    for idx, (date, serial) in enumerate(zip(dates, serials), 1):
        col = idx * 2 - 1
        merge_write(ws, f"{get_column_letter(col)}6:{get_column_letter(col+1)}6", date, BLUE, Font(name="맑은 고딕", bold=True, color=NAVY), Alignment(horizontal="center"))
        merge_write(ws, f"{get_column_letter(col)}7:{get_column_letter(col+1)}7", f"내부 숫자 {serial}", GREEN, Font(name="맑은 고딕", bold=True), Alignment(horizontal="center"))
        for cell in ws[f"{get_column_letter(col)}6:{get_column_letter(col+1)}6"][0]:
            cell.number_format = "yyyy-mm-dd"
    merge_write(ws, "A8:J8", "하루가 지나면 내부 숫자가 1 증가한다 → 그래서 종료일 - 시작일로 기간을 잴 수 있다.", YELLOW, Font(name="맑은 고딕", bold=True, color=NAVY), Alignment(horizontal="center"))
    section(ws, 10, "그림 2. 시간은 하루를 잘게 나눈 소수다")
    timeline = [("00:00", 0), ("06:00", 0.25), ("12:00", 0.5), ("18:00", 0.75), ("24:00", 1)]
    for idx, (label, value) in enumerate(timeline, 1):
        col = idx * 2 - 1
        merge_write(ws, f"{get_column_letter(col)}12:{get_column_letter(col+1)}12", label, BLUE, Font(name="맑은 고딕", bold=True, color=NAVY), Alignment(horizontal="center"))
        merge_write(ws, f"{get_column_letter(col)}13:{get_column_letter(col+1)}13", value, GREEN, Font(name="맑은 고딕", bold=True), Alignment(horizontal="center"))
    note(ws, "A15:J16", "초보자 함정\n날짜처럼 보이게 하는 것은 '표시 형식'입니다. 셀 안에는 숫자가 들어 있습니다. 시간 차이를 구한 뒤 24를 곱하면 '시간 수'가 됩니다.")
    table(
        ws,
        18,
        ["상황", "시작", "종료", "수식", "결과", "쉽게 말하면"],
        [
            ["날짜 차이", datetime(2026, 6, 1), datetime(2026, 6, 10), "=C19-B19", "=C19-B19", "며칠 지났나?"],
            ["시간 차이", time(9, 0), time(17, 30), "=(C20-B20)*24", "=(C20-B20)*24", "몇 시간 일했나?"],
        ],
        formats={2: "yyyy-mm-dd;[Red]-yyyy-mm-dd", 3: "yyyy-mm-dd;[Red]-yyyy-mm-dd", 5: "0.0"},
        fills={4: BLUE, 5: BLUE},
    )
    ws["B20"].number_format = "hh:mm"
    ws["C20"].number_format = "hh:mm"
    ws["E19"].number_format = "0"


def make_duration_sheet(wb):
    ws = wb.create_sheet("0_기간계산공식")
    setup_sheet(ws, [18, 22, 25, 25, 24, 34, 20, 22])
    title(ws, "0. 기간 계산 공식: 먼저 단위를 본다", "근속일수, 근무시간, 경과일은 모두 빼기입니다. 결과 단위 때문에 마지막 처리가 달라집니다.", 8)
    section(ws, 4, "그림 1. 문제 문장을 공식으로 번역하기", 8)
    table(
        ws,
        5,
        ["문제에서 묻는 것", "시작", "끝 또는 기준", "기본 수식", "결과 단위", "왜 이렇게 하나?", "시작 포함?", "시험 단서"],
        [
            ["경과일", "시작일", "종료일", "공식: =종료일-시작일", "일", "날짜는 하루마다 1씩 증가", "포함이면 +1", "경과일, 소요일수"],
            ["근속일수", "입사일", "오늘", "공식: =TODAY()-입사일", "일", "오늘까지 며칠 지났는지 계산", "문제 지시 확인", "근속일수, 재직일수"],
            ["근무시간(숫자)", "출근", "퇴근", "공식: =(퇴근-출근)*24", "시간 수", "시간 차이는 하루 기준 소수라서 24를 곱함", "해당 없음", "몇 시간, 근무시간"],
            ["근무시간(시:분)", "출근", "퇴근", "공식: =퇴근-출근", "시:분", "시간 표시 형식으로 보여주면 24를 곱하지 않음", "해당 없음", "hh:mm 형식"],
        ],
        fills={4: BLUE, 5: YELLOW},
    )
    note(ws, "A11:H12", "한 줄 암기\n날짜끼리 빼면 이미 '일수'다. 시간끼리 빼면 '하루의 일부'다. 몇 시간인지 숫자로 보여 달라면 마지막에 *24를 붙인다.")
    section(ws, 14, "그림 2. 시작일 포함 여부: 문제 문장을 꼭 확인", 8)
    merge_write(ws, "A16:B17", "6월 1일\n시작", BLUE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "C16:F17", "경과한 거리: 9일\n=종료일-시작일\n\n시작일과 종료일을 모두 세기: 10일\n=종료일-시작일+1", YELLOW, Font(name="맑은 고딕", size=11, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "G16:H17", "6월 10일\n종료", GREEN, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    note(ws, "A19:H20", "주의\n그냥 '경과일'이면 종료일-시작일. '시작일을 포함하여', '총 며칠 동안'처럼 양 끝 날짜를 모두 세라고 하면 +1을 붙입니다.")
    section(ws, 22, "따라 쓰는 예제: 결과 단위를 비교", 8)
    table(
        ws,
        23,
        ["종류", "시작 또는 출근", "종료 또는 퇴근", "수식", "결과", "표시 형식", "읽는 법", "핵심"],
        [
            ["경과일", datetime(2026, 6, 1), datetime(2026, 6, 10), "=C24-B24", "=C24-B24", "숫자", "9일", "날짜 차이는 그대로"],
            ["양 끝 포함 일수", datetime(2026, 6, 1), datetime(2026, 6, 10), "=C25-B25+1", "=C25-B25+1", "숫자", "10일", "포함 지시가 있을 때만 +1"],
            ["근속일수", datetime(2020, 1, 1), "=TODAY()", "=TODAY()-B26", "=TODAY()-B26", "숫자", "오늘까지 일수", "오늘-입사일"],
            ["근무시간 숫자", time(9, 0), time(17, 30), "=(C27-B27)*24", "=(C27-B27)*24", "숫자", "8.5시간", "몇 시간이면 *24"],
            ["근무시간 시:분", time(9, 0), time(17, 30), "=C28-B28", "=C28-B28", "[h]:mm", "8:30", "시:분이면 *24 없음"],
        ],
        fills={4: BLUE, 5: BLUE},
    )
    for cell in ["B24", "C24", "B25", "C25", "B26", "C26"]:
        ws[cell].number_format = "yyyy-mm-dd"
    for cell in ["B27", "C27", "B28", "C28"]:
        ws[cell].number_format = "hh:mm"
    ws["E24"].number_format = "0"
    ws["E25"].number_format = "0"
    ws["E26"].number_format = "0"
    ws["E27"].number_format = '0.0"시간"'
    ws["E28"].number_format = "[h]:mm"
    section(ws, 30, "직접 입력 연습: 초록칸을 채우기", 8)
    table(
        ws,
        31,
        ["문제", "시작 또는 출근", "종료 또는 퇴근", "정답 입력", "단위", "힌트", "결과 예상", "확인"],
        [
            ["경과일", datetime(2026, 7, 1), datetime(2026, 7, 15), None, "일", "종료-시작", "14", "□"],
            ["양 끝 포함 일수", datetime(2026, 7, 1), datetime(2026, 7, 15), None, "일", "종료-시작+1", "15", "□"],
            ["근속일수", datetime(2020, 1, 1), "=TODAY()", None, "일", "TODAY()-입사일", "오늘까지 일수", "□"],
            ["근무시간 숫자", time(8, 30), time(18, 0), None, "시간", "(퇴근-출근)*24", "9.5", "□"],
            ["근무시간 시:분", time(8, 30), time(18, 0), None, "시:분", "퇴근-출근", "9:30", "□"],
        ],
        fills={4: GREEN},
    )
    for cell in ["B32", "C32", "B33", "C33", "B34", "C34"]:
        ws[cell].number_format = "yyyy-mm-dd"
    for cell in ["B35", "C35", "B36", "C36"]:
        ws[cell].number_format = "hh:mm"
    ws["D32"].number_format = "0"
    ws["D33"].number_format = "0"
    ws["D34"].number_format = "0"
    ws["D35"].number_format = '0.0"시간"'
    ws["D36"].number_format = "[h]:mm"
    note(ws, "A38:H39", "추가 주의: 자정을 넘기는 야간 근무\n퇴근 시간이 출근 시간보다 작아질 수 있습니다. 그때는 =MOD(퇴근-출근,1)*24를 사용합니다. 기본 문제를 익힌 뒤 적용하세요.")


def make_extract_sheet(wb):
    ws = wb.create_sheet("1교시_오늘과추출")
    setup_sheet(ws, [15, 18, 18, 18, 18, 18, 21, 30])
    title(ws, "1교시. 오늘 날짜와 날짜 조각 꺼내기", "TODAY, NOW, YEAR, MONTH, DAY를 일상 예제로 익힙니다.", 8)
    section(ws, 4, "눈으로 보는 함수 지도", 8)
    merge_write(ws, "A6:B7", "TODAY()\n오늘 날짜", BLUE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "D6:E7", "NOW()\n지금 날짜 + 시간", GREEN, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "G6:H7", "자동 갱신\n파일을 열거나 재계산할 때", YELLOW, Font(name="맑은 고딕", size=11, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "A9:H9", "날짜 한 덩어리 2026-06-01  →  YEAR(날짜)=2026  /  MONTH(날짜)=6  /  DAY(날짜)=1", PURPLE, Font(name="맑은 고딕", bold=True, color=NAVY), Alignment(horizontal="center"))
    section(ws, 11, "따라 쓰는 예제", 8)
    table(
        ws,
        12,
        ["이름", "생년월일", "입사일", "오늘", "출생연도", "입사월", "입사일자", "수식 읽기"],
        [
            ["김하늘", datetime(1997, 8, 15), datetime(2021, 3, 2), "=TODAY()", "=YEAR(B13)", "=MONTH(C13)", "=DAY(C13)", "B13에서 연도만 / C13에서 월만 / 일만 꺼냄"],
            ["이바다", datetime(2000, 1, 7), datetime(2024, 11, 18), "=TODAY()", "=YEAR(B14)", "=MONTH(C14)", "=DAY(C14)", "같은 함수를 아래로 복사"],
            ["박구름", datetime(1989, 12, 29), datetime(2018, 7, 9), "=TODAY()", "=YEAR(B15)", "=MONTH(C15)", "=DAY(C15)", "같은 함수를 아래로 복사"],
        ],
        formats={2: "yyyy-mm-dd", 3: "yyyy-mm-dd", 4: "yyyy-mm-dd"},
        fills={4: BLUE, 5: BLUE, 6: BLUE, 7: BLUE},
    )
    note(ws, "A17:H18", "시험 단서 번역\n'현재 날짜' → TODAY() / '현재 날짜와 시간' → NOW() / '연도만' → YEAR / '월만' → MONTH / '일만' → DAY")
    section(ws, 20, "직접 입력 연습: 초록칸을 채우고 아래로 복사", 8)
    table(
        ws,
        21,
        ["사원", "입사일", "입사연도", "입사월", "입사일자", "오늘", "근속연수", "문제"],
        [
            ["최초롱", datetime(2019, 4, 12), None, None, None, None, None, "C:F는 추출/TODAY, G는 현재연도-입사연도"],
            ["정새벽", datetime(2022, 10, 3), None, None, None, None, None, "22행 수식을 아래로 복사"],
            ["윤여름", datetime(2016, 1, 25), None, None, None, None, None, "22행 수식을 아래로 복사"],
        ],
        formats={2: "yyyy-mm-dd", 6: "yyyy-mm-dd"},
        fills={3: GREEN, 4: GREEN, 5: GREEN, 6: GREEN, 7: GREEN},
    )
    note(ws, "A27:H28", "정답은 '정답해설' 시트에서 확인하세요. TODAY와 NOW는 괄호 안에 아무것도 넣지 않습니다.")


def make_build_time_sheet(wb):
    ws = wb.create_sheet("2교시_만들기와시간")
    setup_sheet(ws, [15, 15, 15, 18, 18, 18, 18, 30])
    title(ws, "2교시. 조각을 합쳐 날짜와 시간 만들기", "DATE, TIME으로 만들고 HOUR, MINUTE, SECOND로 다시 꺼냅니다.", 8)
    section(ws, 4, "그림 1. DATE는 레고처럼 연, 월, 일을 조립한다", 8)
    merge_write(ws, "A6:B7", "2026\n연", BLUE, Font(name="맑은 고딕", size=14, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center"))
    merge_write(ws, "C6:D7", "6\n월", GREEN, Font(name="맑은 고딕", size=14, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center"))
    merge_write(ws, "E6:F7", "15\n일", YELLOW, Font(name="맑은 고딕", size=14, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center"))
    merge_write(ws, "G6:H7", "예시: =DATE(2026,6,15)\n→ 2026-06-15", PURPLE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    section(ws, 9, "그림 2. TIME도 시, 분, 초를 조립한다", 8)
    merge_write(ws, "A11:B12", "9\n시", BLUE, Font(name="맑은 고딕", size=14, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center"))
    merge_write(ws, "C11:D12", "30\n분", GREEN, Font(name="맑은 고딕", size=14, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center"))
    merge_write(ws, "E11:F12", "0\n초", YELLOW, Font(name="맑은 고딕", size=14, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center"))
    merge_write(ws, "G11:H12", "예시: =TIME(9,30,0)\n→ 09:30:00", PURPLE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    section(ws, 14, "따라 쓰는 예제", 8)
    table(
        ws,
        15,
        ["연", "월", "일", "DATE 결과", "시", "분", "TIME 결과", "꺼내기"],
        [
            [2026, 6, 15, "=DATE(A16,B16,C16)", 9, 30, "=TIME(E16,F16,0)", "=HOUR(G16)&\"시 \"&MINUTE(G16)&\"분\""],
            [2026, 12, 31, "=DATE(A17,B17,C17)", 14, 5, "=TIME(E17,F17,0)", "=HOUR(G17)&\"시 \"&MINUTE(G17)&\"분\""],
        ],
        formats={4: "yyyy-mm-dd", 7: "hh:mm"},
        fills={4: BLUE, 7: BLUE, 8: BLUE},
    )
    note(ws, "A19:H20", "시험 단서 번역\n'연, 월, 일을 이용하여 날짜를 표시' → DATE / '시, 분, 초를 이용하여 시간 표시' → TIME / '시간에서 시만 추출' → HOUR")
    section(ws, 22, "직접 입력 연습: 초록칸을 채우기", 8)
    table(
        ws,
        23,
        ["연", "월", "일", "조립한 날짜", "출근", "퇴근", "근무시간", "문제"],
        [
            [2026, 7, 20, None, time(9, 0), time(18, 0), None, "D는 DATE, G는 (퇴근-출근)*24"],
            [2026, 9, 3, None, time(8, 30), time(17, 30), None, "24행 수식을 아래로 복사"],
            [2027, 1, 1, None, time(10, 0), time(16, 30), None, "24행 수식을 아래로 복사"],
        ],
        formats={4: "yyyy-mm-dd", 5: "hh:mm", 6: "hh:mm", 7: '0.0"시간"'},
        fills={4: GREEN, 7: GREEN},
    )


def make_weekday_days_sheet(wb):
    ws = wb.create_sheet("3교시_요일과기간")
    setup_sheet(ws, [16, 18, 18, 18, 18, 20, 22, 30])
    title(ws, "3교시. 무슨 요일인지, 며칠 걸렸는지 구하기", "처음에는 공식 두 개만 따라 하세요. 어려운 말은 나중에 봐도 됩니다.", 8)
    note(ws, "A4:H5", "오늘 배울 것은 딱 두 개입니다.\n1. 무슨 요일인지 숫자로 표시하기: =WEEKDAY(날짜,2)  /  2. 며칠 차이인지 구하기: =끝나는날-시작하는날")
    section(ws, 7, "1. 무슨 요일인가? 처음에는 이것 하나만 따라 쓰기", 8)
    merge_write(ws, "A9:D10", "날짜가 C38에 있다면\n=WEEKDAY(C38,2)", BLUE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "E9:H10", "뒤의 2는 이렇게 약속한다는 뜻\n월=1, 화=2, 수=3, 목=4, 금=5, 토=6, 일=7", GREEN, Font(name="맑은 고딕", size=11, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    note(ws, "A12:H13", "왜 요일 이름 대신 숫자가 나오나요?\n시험 문제에서 '월요일은 1, 화요일은 2처럼 표시하시오'라고 시키기 때문입니다. 일단 =WEEKDAY(날짜셀,2)를 그대로 따라 쓰면 됩니다.")
    section(ws, 15, "참고: 시험 문제가 다르게 시킬 때만 아래 표 보기", 8)
    table(
        ws, 16,
        ["문제의 말", "월", "화", "수", "목", "금", "토", "일"],
        [
            ["월요일을 1로 표시", 1, 2, 3, 4, 5, 6, 7],
            ["일요일을 1로 표시", 2, 3, 4, 5, 6, 7, 1],
        ],
        fills={1: YELLOW},
    )
    note(ws, "A20:H21", "일요일을 1로 하라고 할 때만 =WEEKDAY(날짜셀,1)을 사용합니다. 문제에 특별한 말이 없다면 기본 수업에서는 월요일=1 방식부터 연습합니다.")
    section(ws, 23, "2. 며칠 걸렸나? 끝나는 날에서 시작하는 날 빼기", 8)
    merge_write(ws, "A25:B26", "6월 1일\n시작", BLUE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "C25:F26", "→ 9일 지남 →\n=끝나는날-시작하는날", YELLOW, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "G25:H26", "6월 10일\n끝", GREEN, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    note(ws, "A28:H29", "예: 시작일이 B32, 종료일이 C32라면 =C32-B32\n주의: 시작일과 종료일을 모두 날짜 수에 포함하라고 하면 마지막에 +1을 붙입니다.")
    section(ws, 31, "3. 완성 예제를 보고 그대로 따라 하기", 8)
    table(
        ws,
        32,
        ["업무", "시작일", "종료일", "며칠 걸림", "끝나는 날 요일", "일수 공식", "요일 공식", "읽는 순서"],
        [
            ["교육", datetime(2026, 6, 1), datetime(2026, 6, 10), "=C33-B33", "=WEEKDAY(C33,2)", "=C33-B33", "=WEEKDAY(C33,2)", "끝-시작 / 끝나는 날의 요일"],
            ["점검", datetime(2026, 7, 13), datetime(2026, 7, 17), "=C34-B34", "=WEEKDAY(C34,2)", "=C34-B34", "=WEEKDAY(C34,2)", "첫 줄을 만든 뒤 아래로 복사"],
        ],
        formats={2: "yyyy-mm-dd", 3: "yyyy-mm-dd"},
        fills={4: BLUE, 5: BLUE},
    )
    section(ws, 36, "4. 직접 입력: 초록칸 두 개부터 채우기", 8)
    table(
        ws,
        37,
        ["계약", "시작일", "종료일", "며칠 걸림", "끝나는 날 요일", "끝나는 달", "끝나는 날짜", "처음 할 일"],
        [
            ["A-101", datetime(2026, 8, 3), datetime(2026, 8, 28), None, None, None, None, "먼저 D38에 =C38-B38"],
            ["A-102", datetime(2026, 9, 14), datetime(2026, 10, 2), None, None, None, None, "다음 E38에 =WEEKDAY(C38,2)"],
            ["A-103", datetime(2026, 11, 2), datetime(2026, 11, 30), None, None, None, None, "F38은 =MONTH(C38), G38은 =DAY(C38)"],
        ],
        formats={2: "yyyy-mm-dd", 3: "yyyy-mm-dd"},
        fills={4: GREEN, 5: GREEN, 6: GREEN, 7: GREEN},
    )
    note(ws, "A42:H43", "막히면 한 번에 네 칸을 풀지 마세요.\nD38 하나만 입력 → 값이 나오면 E38 입력 → F38 입력 → G38 입력 → 마지막에 아래로 복사합니다.")


def make_due_sheet(wb):
    ws = wb.create_sheet("4교시_납기와월말")
    setup_sheet(ws, [18, 18, 18, 20, 20, 20, 22, 28])
    title(ws, "4교시. 주말을 건너뛰고, 몇 개월 뒤와 월말 찾기", "WORKDAY, EDATE, EOMONTH를 납기일과 만기일 문제에 연결합니다.", 8)
    section(ws, 4, "그림. 함수 세 개를 상황별로 고르기", 8)
    merge_write(ws, "A6:B8", "WORKDAY\n주말을 건너뜀\n납기일", BLUE, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "D6:E8", "EDATE\n몇 개월 뒤\n만기일", GREEN, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    merge_write(ws, "G6:H8", "EOMONTH\n해당 월의 마지막 날\n월말", YELLOW, Font(name="맑은 고딕", size=12, bold=True, color=NAVY), Alignment(horizontal="center", vertical="center", wrap_text=True))
    note(ws, "A10:H11", "외우는 문장\nWORKDAY는 주말을 피해 걷는다. EDATE는 달력에서 몇 장 넘긴다. EOMONTH는 넘긴 달력의 맨 끝 칸으로 간다.")
    section(ws, 13, "따라 쓰는 예제", 8)
    table(
        ws,
        14,
        ["업무", "기준일", "근무일수", "개월수", "납기일", "몇개월뒤", "해당월말", "설명"],
        [
            ["디자인", datetime(2026, 6, 5), 5, 3, "=WORKDAY(B15,C15)", "=EDATE(B15,D15)", "=EOMONTH(B15,D15)", "휴일 목록이 없으면 주말만 제외"],
            ["검수", datetime(2026, 11, 27), 4, 1, "=WORKDAY(B16,C16)", "=EDATE(B16,D16)", "=EOMONTH(B16,D16)", "아래로 복사"],
        ],
        formats={2: "yyyy-mm-dd", 5: "yyyy-mm-dd", 6: "yyyy-mm-dd", 7: "yyyy-mm-dd"},
        fills={5: BLUE, 6: BLUE, 7: BLUE},
    )
    note(ws, "A18:H19", "조금 더 실전처럼\n휴일 목록까지 제외하라는 문제라면 =WORKDAY(기준일, 근무일수, 휴일범위)처럼 세 번째 인수를 넣습니다.")
    section(ws, 21, "직접 입력 연습: 초록칸을 채우기", 8)
    table(
        ws,
        22,
        ["주문", "주문일", "작업일", "개월", "납기일", "보증만료일", "만료월말", "문제"],
        [
            ["O-201", datetime(2026, 6, 12), 7, 6, None, None, None, "E는 WORKDAY, F는 EDATE, G는 EOMONTH(F,0)"],
            ["O-202", datetime(2026, 8, 28), 3, 12, None, None, None, "23행 수식을 아래로 복사"],
            ["O-203", datetime(2026, 12, 24), 5, 3, None, None, None, "23행 수식을 아래로 복사"],
        ],
        formats={2: "yyyy-mm-dd", 5: "yyyy-mm-dd", 6: "yyyy-mm-dd", 7: "yyyy-mm-dd"},
        fills={5: GREEN, 6: GREEN, 7: GREEN},
    )


def make_exam_sheet(wb):
    ws = wb.create_sheet("실전연습")
    setup_sheet(ws, [14, 17, 16, 16, 16, 18, 18, 18, 18, 20, 32])
    title(ws, "실전연습. 컴활 2급 실기형 날짜·시간 함수", "초록칸에 수식을 직접 입력하고 아래로 복사합니다. 문제 문장에서 단서를 먼저 동그라미 치세요.", 11)
    note(ws, "A4:K5", "시험 풀이 순서\n1. 결과 칸이 무엇을 원하는지 읽기 → 2. 날짜를 꺼낼지, 만들지, 거리를 잴지 결정 → 3. 첫 행에 수식 입력 → 4. 채우기 핸들로 아래 행 복사")
    section(ws, 7, "세트 A. 회원 관리표", 11)
    table(
        ws,
        8,
        ["회원번호", "가입일", "생년월일", "개월", "가입연도", "가입월", "생년", "개월후", "개월후월말", "가입요일(월=1)", "문제"],
        [
            ["M-001", datetime(2024, 1, 17), datetime(1995, 5, 2), 6, None, None, None, None, None, None, "E:J열 수식 작성"],
            ["M-002", datetime(2025, 11, 28), datetime(2001, 12, 24), 12, None, None, None, None, None, None, "9행 수식 아래 복사"],
            ["M-003", datetime(2026, 3, 9), datetime(1988, 7, 11), 3, None, None, None, None, None, None, "9행 수식 아래 복사"],
            ["M-004", datetime(2026, 5, 25), datetime(1999, 10, 3), 1, None, None, None, None, None, None, "9행 수식 아래 복사"],
        ],
        formats={2: "yyyy-mm-dd", 3: "yyyy-mm-dd", 8: "yyyy-mm-dd", 9: "yyyy-mm-dd"},
        fills={5: GREEN, 6: GREEN, 7: GREEN, 8: GREEN, 9: GREEN, 10: GREEN},
    )
    section(ws, 15, "세트 B. 주문 납기표", 11)
    table(
        ws,
        16,
        ["주문번호", "주문일", "작업일", "출근", "퇴근", "납기일", "근무시간", "납기연도", "납기월", "납기일자", "문제"],
        [
            ["P-101", datetime(2026, 6, 5), 5, time(9, 0), time(17, 30), None, None, None, None, None, "F:J열 수식 작성"],
            ["P-102", datetime(2026, 7, 17), 7, time(8, 30), time(18, 0), None, None, None, None, None, "17행 수식 아래 복사"],
            ["P-103", datetime(2026, 9, 25), 3, time(10, 0), time(16, 0), None, None, None, None, None, "17행 수식 아래 복사"],
            ["P-104", datetime(2026, 12, 24), 4, time(9, 30), time(18, 30), None, None, None, None, None, "17행 수식 아래 복사"],
        ],
        formats={2: "yyyy-mm-dd", 4: "hh:mm", 5: "hh:mm", 6: "yyyy-mm-dd", 7: '0.0"시간"'},
        fills={6: GREEN, 7: GREEN, 8: GREEN, 9: GREEN, 10: GREEN},
    )
    section(ws, 23, "세트 C. 계약 관리표", 11)
    table(
        ws,
        24,
        ["계약번호", "시작일", "종료일", "기간", "종료요일(일=1)", "현재일", "현재시각", "종료월말", "종료연도", "종료월", "문제"],
        [
            ["C-01", datetime(2026, 1, 2), datetime(2026, 6, 30), None, None, None, None, None, None, None, "D:J열 수식 작성"],
            ["C-02", datetime(2026, 2, 16), datetime(2026, 8, 14), None, None, None, None, None, None, None, "25행 수식 아래 복사"],
            ["C-03", datetime(2026, 5, 11), datetime(2026, 12, 18), None, None, None, None, None, None, None, "25행 수식 아래 복사"],
        ],
        formats={2: "yyyy-mm-dd", 3: "yyyy-mm-dd", 6: "yyyy-mm-dd", 7: "yyyy-mm-dd hh:mm", 8: "yyyy-mm-dd"},
        fills={4: GREEN, 5: GREEN, 6: GREEN, 7: GREEN, 8: GREEN, 9: GREEN, 10: GREEN},
    )
    note(ws, "A30:K31", "채점 기준\n함수 이름, 인수 순서, 괄호, 쉼표를 확인합니다. 날짜 차이는 종료일-시작일 순서입니다. 초록칸 65개 중 52개 이상이면 다음 단원으로 이동하세요.")


def make_answers(wb):
    ws = wb.create_sheet("정답해설")
    setup_sheet(ws, [22, 14, 34, 58])
    title(ws, "정답해설", "학생 풀이 후 확인용입니다. 수식과 생각 순서를 같이 봅니다.", 4)
    rows = [
        ["0_기간계산공식", "D32", "=C32-B32", "날짜끼리 빼면 결과는 이미 일수입니다."],
        ["0_기간계산공식", "D33", "=C33-B33+1", "시작일과 종료일을 모두 세라는 문제이므로 +1을 붙입니다."],
        ["0_기간계산공식", "D34", "=TODAY()-B34", "오늘에서 입사일을 빼면 오늘까지의 근속일수가 나옵니다."],
        ["0_기간계산공식", "D35", "=(C35-B35)*24", "몇 시간인지 숫자로 표시하므로 시간 차이에 24를 곱합니다."],
        ["0_기간계산공식", "D36", "=C36-B36", "시:분 형식으로 표시하므로 24를 곱하지 않습니다."],
        ["1교시_오늘과추출", "C22", "=YEAR(B22)", "입사일 B22에서 연도만 꺼냅니다."],
        ["1교시_오늘과추출", "D22", "=MONTH(B22)", "입사일 B22에서 월만 꺼냅니다."],
        ["1교시_오늘과추출", "E22", "=DAY(B22)", "입사일 B22에서 일만 꺼냅니다."],
        ["1교시_오늘과추출", "F22", "=TODAY()", "현재 날짜 함수는 인수가 없습니다."],
        ["1교시_오늘과추출", "G22", "=YEAR(TODAY())-YEAR(B22)", "현재연도에서 입사연도를 뺍니다. 시험형 단순 근속연수 계산입니다."],
        ["2교시_만들기와시간", "D24", "=DATE(A24,B24,C24)", "연, 월, 일 조각을 순서대로 넣어 날짜를 만듭니다."],
        ["2교시_만들기와시간", "G24", "=(F24-E24)*24", "시간은 하루의 일부이므로 차이에 24를 곱해 시간 수로 만듭니다."],
        ["3교시_요일과기간", "D38", "=C38-B38", "끝나는 날에서 시작하는 날을 빼면 며칠 걸렸는지 나옵니다."],
        ["3교시_요일과기간", "E38", "=WEEKDAY(C38,2)", "월요일을 1로 표시하는 약속입니다. 처음에는 이 모양 그대로 따라 쓰세요."],
        ["3교시_요일과기간", "F38", "=MONTH(C38)", "끝나는 날에서 달 숫자만 꺼냅니다."],
        ["3교시_요일과기간", "G38", "=DAY(C38)", "끝나는 날에서 날짜 숫자만 꺼냅니다."],
        ["4교시_납기와월말", "E23", "=WORKDAY(B23,C23)", "주문일에서 작업일 수만큼 주말을 건너뜁니다."],
        ["4교시_납기와월말", "F23", "=EDATE(B23,D23)", "주문일에서 지정 개월 수 뒤 날짜입니다."],
        ["4교시_납기와월말", "G23", "=EOMONTH(F23,0)", "보증만료일이 들어 있는 달의 마지막 날입니다."],
        ["실전연습 A", "E9", "=YEAR(B9)", "가입일의 연도를 추출합니다."],
        ["실전연습 A", "F9", "=MONTH(B9)", "가입일의 월을 추출합니다."],
        ["실전연습 A", "G9", "=YEAR(C9)", "생년월일의 연도를 추출합니다."],
        ["실전연습 A", "H9", "=EDATE(B9,D9)", "가입일에서 D9개월 뒤 날짜입니다."],
        ["실전연습 A", "I9", "=EOMONTH(H9,0)", "H9 날짜가 포함된 달의 월말입니다."],
        ["실전연습 A", "J9", "=WEEKDAY(B9,2)", "월요일을 1로 계산합니다."],
        ["실전연습 B", "F17", "=WORKDAY(B17,C17)", "주문일부터 작업일 수만큼 주말을 제외합니다."],
        ["실전연습 B", "G17", "=(E17-D17)*24", "퇴근-출근 후 24를 곱합니다."],
        ["실전연습 B", "H17", "=YEAR(F17)", "납기일의 연도를 추출합니다."],
        ["실전연습 B", "I17", "=MONTH(F17)", "납기일의 월을 추출합니다."],
        ["실전연습 B", "J17", "=DAY(F17)", "납기일의 일을 추출합니다."],
        ["실전연습 C", "D25", "=C25-B25", "종료일에서 시작일을 빼서 두 날짜 사이 거리를 계산합니다."],
        ["실전연습 C", "E25", "=WEEKDAY(C25)", "일요일을 1로 쓰는 기본 번호표이므로 두 번째 인수를 생략할 수 있습니다."],
        ["실전연습 C", "F25", "=TODAY()", "현재 날짜입니다."],
        ["실전연습 C", "G25", "=NOW()", "현재 날짜와 시간입니다."],
        ["실전연습 C", "H25", "=EOMONTH(C25,0)", "종료일이 속한 달의 마지막 날짜입니다."],
        ["실전연습 C", "I25", "=YEAR(C25)", "종료일의 연도입니다."],
        ["실전연습 C", "J25", "=MONTH(C25)", "종료일의 월입니다."],
    ]
    table(ws, 5, ["시트", "첫 입력 셀", "정답 수식", "왜 이렇게 쓰나?"], rows, fills={3: PURPLE})
    for row in range(6, 6 + len(rows)):
        ws.cell(row, 3).value = f"입력: {ws.cell(row, 3).value}"
    note(ws, f"A{len(rows)+8}:D{len(rows)+9}", "아래로 복사할 때 셀 주소가 행마다 바뀌는지 확인하세요. TODAY(), NOW()처럼 빈 괄호가 필요한 함수도 있습니다.")


def make_script_sheet(wb):
    ws = wb.create_sheet("강사용대본")
    setup_sheet(ws, [16, 22, 62, 36])
    title(ws, "강사용대본", "4시간 수업 운영용입니다. 설명 뒤에는 반드시 학생이 손으로 입력하게 합니다.", 4)
    rows = [
        ["0:00-0:10", "도입", "날짜 함수는 달력을 다루는 도구입니다. 외울 함수가 많아 보여도 하는 일은 꺼내기, 만들기, 거리 재기 세 가지입니다.", "세 가지를 학생이 말하게 하기"],
        ["0:10-0:20", "숫자 원리", "엑셀은 날짜를 숫자로 저장합니다. 하루가 지나면 1이 늘어납니다. 시간은 하루를 잘게 나눈 소수입니다.", "날짜 차이와 시간 차이 예제 확인"],
        ["0:20-0:40", "기간 공식", "날짜끼리 빼면 일수입니다. 시간끼리 빼면 하루의 일부이므로 몇 시간인지 숫자로 볼 때만 24를 곱합니다. 포함 지시가 있으면 날짜 차이에 1을 더합니다.", "기간계산공식 초록칸 입력"],
        ["0:40-1:00", "TODAY/NOW", "TODAY는 오늘 달력 한 장, NOW는 오늘 달력에 시계까지 붙인 것입니다. 둘 다 빈 괄호입니다.", "TODAY(), NOW() 직접 입력"],
        ["1:00-1:15", "YEAR/MONTH/DAY", "큰 날짜 덩어리에서 필요한 조각만 꺼냅니다. 주민등록 생년, 입사월 같은 문제를 연결합니다.", "1교시 초록칸 입력"],
        ["1:15-1:35", "DATE/TIME", "반대로 조각을 합칠 때는 DATE와 TIME입니다. 순서는 연월일, 시분초입니다.", "조립 그림을 손가락으로 따라가기"],
        ["1:35-1:55", "시간 차이", "시간끼리 빼면 하루 기준 소수가 나옵니다. 몇 시간인지 묻는다면 24를 곱합니다.", "2교시 초록칸 입력"],
        ["1:55-2:15", "WEEKDAY", "요일 함수는 번호표를 고르는 문제입니다. 월요일을 1로 하라는 단서를 놓치지 않습니다.", "번호표 1과 2 비교"],
        ["2:15-2:35", "날짜 빼기", "거리 재기입니다. 종료일에서 시작일을 빼면 며칠 차이인지 나옵니다. DAYS 함수가 없는 Excel에서도 이 방식은 동작합니다.", "3교시 초록칸 입력"],
        ["2:35-3:00", "WORKDAY", "납기일인데 주말을 쉬어야 한다면 WORKDAY입니다. 휴일 목록이 주어지면 세 번째 인수를 추가합니다.", "주말이 끼는 예제 확인"],
        ["3:00-3:20", "EDATE/EOMONTH", "몇 개월 뒤 같은 날짜는 EDATE, 그 달의 마지막 날은 EOMONTH입니다.", "4교시 초록칸 입력"],
        ["3:20-3:50", "실전", "함수를 바로 쓰지 말고 문제 문장에서 단서를 먼저 찾게 합니다. 첫 행만 완성한 뒤 아래로 복사합니다.", "세트 A/B/C 풀이"],
        ["3:50-4:00", "오답 교정", "틀린 문제는 함수 이름보다 '왜 그 함수를 골랐는지'를 다시 말하게 합니다.", "52개 이상 정답 여부 확인"],
    ]
    table(ws, 5, ["시간", "진행", "강사 멘트", "학생 확인"], rows)
    note(ws, "A19:D20", "강사 메모\n단순 근속연수 =YEAR(TODAY())-YEAR(입사일)은 시험형 연도 차이 예제입니다. 정확한 만 나이나 정확한 근속기간이 필요하면 생일/기념일 경과 여부까지 따져야 합니다.")


def finalize(wb):
    for ws in wb.worksheets:
        style_all(ws)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    alignment = copy(cell.alignment)
                    alignment.vertical = alignment.vertical or "center"
                    cell.alignment = alignment
        ws.auto_filter.ref = None
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.properties.title = "컴활 2급 실기 - 날짜와 시간 함수 4시간 학습지"
    wb.properties.subject = "컴퓨터활용능력 2급 실기 날짜와 시간 함수"
    wb.properties.creator = "Codex"
    wb.save(OUTPUT)


def validate():
    wb = load_workbook(OUTPUT, data_only=False)
    required = [
        "표지_4시간계획",
        "0_날짜는숫자",
        "0_기간계산공식",
        "1교시_오늘과추출",
        "2교시_만들기와시간",
        "3교시_요일과기간",
        "4교시_납기와월말",
        "실전연습",
        "정답해설",
        "강사용대본",
    ]
    assert wb.sheetnames == required
    assert wb["1교시_오늘과추출"]["D13"].value == "=TODAY()"
    assert wb["3교시_요일과기간"]["D33"].value == "=C33-B33"
    assert wb["4교시_납기와월말"]["E15"].value == "=WORKDAY(B15,C15)"
    assert wb["실전연습"]["E9"].value is None
    assert wb["정답해설"]["C6"].value == "입력: =C32-B32"
    assert wb["정답해설"]["C11"].value == "입력: =YEAR(B22)"
    print(f"created={OUTPUT}")
    print(f"size={OUTPUT.stat().st_size}")
    for ws in wb.worksheets:
        formulas = sum(
            1 for row in ws.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        print(f"{ws.title}: rows={ws.max_row}, cols={ws.max_column}, formulas={formulas}")


if __name__ == "__main__":
    workbook = Workbook()
    make_cover(workbook)
    make_serial_sheet(workbook)
    make_duration_sheet(workbook)
    make_extract_sheet(workbook)
    make_build_time_sheet(workbook)
    make_weekday_days_sheet(workbook)
    make_due_sheet(workbook)
    make_exam_sheet(workbook)
    make_answers(workbook)
    make_script_sheet(workbook)
    finalize(workbook)
    validate()
